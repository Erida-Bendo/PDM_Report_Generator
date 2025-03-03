from openpyxl import load_workbook
from latex_content import LatexContent as lc
from latexcompiler import LC as compiler


def extract_excel_data(file_path, sheet_column, ag_idx, n_idx, al_idx, ao_idx, ar_idx, as_idx, rend_idx=None):
    
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)    
    # Access the specific sheet
    sheet = wb[sheet_column]
    
    # Initialize list to store results
    results = []
    
    # Define column indices (0-based)
    col_ag_idx = ag_idx  # Column AG
    col_n_idx = n_idx   # Column N
    col_al_idx = al_idx   # Column AL
    col_ao_idx = ao_idx   # Column AO
    col_ar_idx = ar_idx  # Column AR
    col_as_idx = as_idx  # Column AS
    
    row_count = 0
    processed_count = 0
    rows = list(sheet.rows)
    
    for row in sheet.iter_rows(min_row=17):
        row_count += 1
        
        try:
            # Get AG cell value first
            ag_value = row[col_ag_idx].value
            
            
            # If AG column has a value, collect the data
            if ag_value:
                processed_count += 1
                date = rows[11][col_ag_idx].value
                date = str(str(date)[:10])
                dates=[]
                revText=[]
                n_value = row[col_n_idx].value
                al_value = row[col_al_idx].value
                ao_value = row[col_ao_idx].value
                ar_value = row[col_ar_idx].value
                as_value = row[col_as_idx].value

                
                if(rend_idx):
                    for i in range(col_ag_idx-rend_idx):
                        
                        if(row[rend_idx+i].value):
                            revText.append(row[rend_idx+i].value.replace('_','\\_'))
                            value = rows[11][rend_idx+i].value
                            sliced_value = str(str(value)[:10])
                            dates.append(sliced_value)
                        
                        
                
                results.append({
                    "proj_name": n_value,
                    "page_type": al_value,
                    "proj_code": ao_value,
                    "author": ar_value,
                    "checker": as_value,
                    "rev_num": ag_value,
                    "revDates": dates,
                    "revText": revText,
                    "date": date
                })
                
        except IndexError as e:
            print(f"Row {row_count + 16} doesn't have enough columns. Available columns: {len(row)}")
            continue
        except Exception as e:
            print(f"Error processing row {row_count + 16}: {str(e)}")
            continue

    wb.close()
    
    return results

def create_report(disclaimer, language, logopath, date_number, date_month, project_name, pcode,output_folder, output_file,report_name,proj_code,rev_num,author, description, checker,auth_init, check_init, author_email,page_type, revDates, revText, revDate):
    with open(output_file, "a", encoding="utf-8") as tex_file:
       tex_file.write(lc.content(disclaimer, language, logopath, date_number, date_month, project_name,report_name,proj_code, pcode ,rev_num, author, description, checker, auth_init, check_init, author_email,page_type, revDates, revText, revDate))
    compiler.compile_document(tex_engine='lualatex',
                            bib_engine='biber',
                            no_bib=True,
                            path=output_file, folder_name= output_folder)

    print("Done!")   

    return None



